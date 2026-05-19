# excel_export.py
import pandas as pd     #Utilisé pour structurer les données brutes sous forme de tableaux 
import io  #Permet à l'utilisateur de télécharger directement depuis Streamlit
from datetime import datetime  #Utilisé pour capturer la date du jour et l'horodatage
import numpy as np  #Utilisé pour extraire les métriques statistiques des résultats bruts
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side   #Utilisé pour le design du tableau Excel
from openpyxl.utils import get_column_letter   #Utilisé pour pour convertir les index numériques des colonnes en lettres alphabétiques pour ajuster leur largeur

def generate_excel_report(sim_data, df_sens, scenario_name):    #Génération d'un rapport Excel complet
    results = sim_data["annual_loss"]
    lef = sim_data["lef"]
    
    #Préparation des données
    s = {
        "scenario": scenario_name,
        "n_simulations": len(results),
        "ale": results.mean(),
        "median": np.median(results),
        "std": results.std(),
        "var_80": np.percentile(results, 80),
        "var_90": np.percentile(results, 90),
        "var_95": np.percentile(results, 95),
        "var_99": np.percentile(results, 99),
        "prob_zero": (lef == 0).mean(),
        "prob_one": (lef == 1).mean(),
        "prob_two_plus": (lef >= 2).mean(),
    }

    rows_synthese = [
        ("IDENTIFICATION", ""),
        ("Scénario", s["scenario"]),
        ("Date d'analyse", datetime.now().strftime("%d/%m/%Y")),
        ("Simulations", f"{s['n_simulations']:,}"),
        ("", ""),
        ("PERTE ANNUELLE", ""),
        ("ALE — Moyenne", f"{s['ale']:,.0f} €"),
        ("Médiane", f"{s['median']:,.0f} €"),
        ("Écart-type", f"{s['std']:,.0f} €"),
        ("", ""),
        ("VALUE AT RISK", ""),
        ("VaR 80%", f"{s['var_80']:,.0f} €"),
        ("VaR 90%", f"{s['var_90']:,.0f} €"),
        ("VaR 95%", f"{s['var_95']:,.0f} €"),
        ("VaR 99%", f"{s['var_99']:,.0f} €"),
        ("", ""),
        ("FRÉQUENCE ANNUELLE", ""),
        ("Années sans incident", f"{s['prob_zero']:.1%}"),
        ("Années avec 1 incident", f"{s['prob_one']:.1%}"),
        ("Années avec 2+ incidents", f"{s['prob_two_plus']:.1%}"),
    ]
    df_synthese = pd.DataFrame(rows_synthese, columns=["Indicateur", "Valeur"])

    percentiles = list(range(5, 100, 5)) + [99]
    df_percentiles = pd.DataFrame({
        "Percentile": [f"P{p}" for p in percentiles],
        "Perte annuelle (€)": [f"{np.percentile(results, p):,.0f} €" for p in percentiles],
    })

    #Création du fichier Excel
    buffer = io.BytesIO()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2980B9")
    section_font = Font(bold=True, color="FFFFFF", size=10)
    section_fill = PatternFill("solid", fgColor="1A252F")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                         top=Side(style="thin"), bottom=Side(style="thin"))

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_synthese.to_excel(writer, sheet_name="Synthèse", index=False)
        df_sens.to_excel(writer, sheet_name="Analyse sensibilité", index=False)
        df_percentiles.to_excel(writer, sheet_name="Percentiles", index=False)
        
        #Formatage visuel
        wb = writer.book
        for sheet_name, widths in [("Synthèse", [32, 24]), 
                                   ("Analyse sensibilité", [30, 22]), 
                                   ("Percentiles", [14, 22])]:
            ws = wb[sheet_name]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            for cell in ws[1]:
                cell.font, cell.fill, cell.alignment = header_font, header_fill, Alignment(horizontal="center")
            
            if sheet_name == "Synthèse":
                sections = ["IDENTIFICATION", "PERTE ANNUELLE", "VALUE AT RISK", "FRÉQUENCE ANNUELLE"]
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = thin_border
                        if cell.value in sections:
                            cell.font, cell.fill, cell.alignment = section_font, section_fill, Alignment(horizontal="center")

    buffer.seek(0)
    return buffer